# Formatar um relatório (cores, bordas, ajuste de colunas)
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

def formatar_relatorio(arquivo_entrada, arquivo_saida):
    """Aplica formatação profissional: cabeçalho em negrito/azul, bordas, ajuste de coluna."""
    wb = load_workbook(arquivo_entrada)
    ws = wb.active
    
    # Estilos
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Aplica no cabeçalho (primeira linha)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align
    
    # Bordas e alinhamento para todas as células com dados
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
    
    # Ajusta largura das colunas automaticamente
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    wb.save(arquivo_saida)
    print(f"Formatação aplicada: {arquivo_saida}")

formatar_relatorio("dados_brutos.xlsx", "relatorio_formatado.xlsx")
